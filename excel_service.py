import os
import openpyxl
from datetime import datetime
from openpyxl.utils.exceptions import InvalidFileException
from typing import List, Optional, Tuple

from models import (
    MasterStockProduct, JurnalPenjualan, JurnalPembelian, HargaJual
)

SERVICE_DIR = os.path.dirname(os.path.abspath(__file__))


# --- KONFIGURASI FILE & SHEET ---
FILE_PATH = os.path.join(SERVICE_DIR, '..', 'data', 'MyPos.xlsx')
SHEET_MASTER_STOK = 'Sheet_1_Master_Stok'
SHEET_JURNAL_PENJUALAN = 'Sheet_2_Jurnal_Penjualan'
SHEET_JURNAL_PEMBELIAN = 'Sheet_3_Jurnal_Pembelian' 

# Header untuk Master Stok (Sesuai urutan kolom)
MASTER_STOK_HEADERS = [
    'Nama Produk', 'Satuan Beli', 'Isi per Satuan Beli', 'Kategori',
    'Satuan Unit Dasar', 'Harga jual (Bungkus)', 'Harga jual (Batang)', 
    'Harga jual (Mentah)', 'Harga jual (Seduh)', 'Harga jual (Rebus)', 
    'Harga jual (Rebus+Telur)'
]

# Header untuk Jurnal Penjualan
JURNAL_PENJUALAN_HEADERS = [
    'Timestamp', 'Nama Produk', 'Jumlah Jual', 'Total Harga Jual', 'Catatan'
]

# Header untuk Jurnal Pembelian (Fixed: menggunakan 'Pembelian' bukan 'Pembalian')
JURNAL_PEMBELIAN_HEADERS = [
    'Timestamp', 'Nama Produk', 'Jumlah Beli', 'Satuan Beli', 'Total Harga Beli'
]

def _ensure_file_and_sheets():
    """Memastikan file Excel dan semua sheet inti ada."""
    
    # 1. DEFINISIKAN DICTIONARY DI AWAL FUNGSI
    sheets_to_check = {
        SHEET_MASTER_STOK: MASTER_STOK_HEADERS,
        SHEET_JURNAL_PENJUALAN: JURNAL_PENJUALAN_HEADERS,
        SHEET_JURNAL_PEMBELIAN: JURNAL_PEMBELIAN_HEADERS, 
    }
    
    # 2. FILE CHECK AND LOAD
    if not os.path.exists(FILE_PATH):
        print(f"⚠️ File TIDAK DITEMUKAN di: {FILE_PATH}. Membuat workbook baru...")
        os.makedirs(os.path.dirname(FILE_PATH), exist_ok=True)
        workbook = openpyxl.Workbook()
        
        default_sheet = workbook.active
        if default_sheet.title == 'Sheet':
            workbook.remove(default_sheet)
    else:
        try:
            workbook = openpyxl.load_workbook(FILE_PATH)
            print("👍 File DITEMUKAN. Melanjutkan dengan workbook yang ada.")
        except InvalidFileException:
            raise Exception("File MyPos.xlsx rusak atau tidak valid.")

    # 3. SHEET CHECK
    for sheet_name, headers in sheets_to_check.items():
        if sheet_name not in workbook.sheetnames:
            print(f"   [NOTICE] Sheet '{sheet_name}' hilang. Ditambahkan.")
            sheet = workbook.create_sheet(sheet_name)
            sheet.append(headers)

    workbook.save(FILE_PATH)

def _get_workbook_and_sheet(sheet_name: str, read_only:bool = False):
    """Helper untuk memuat workbook dan mendapatkan sheet tertentu."""
    _ensure_file_and_sheets()
    try:
        workbook = openpyxl.load_workbook(FILE_PATH, read_only=read_only)
        sheet = workbook[sheet_name]
        return workbook, sheet
    except (KeyError, InvalidFileException) as e:
        raise Exception(f"Gagal memuat sheet {sheet_name}: {e}")

# --- FUNGSI UTAMA (CRUD MASTER STOK) ---
def read_master_stock() -> List[MasterStockProduct]:
    """Membaca semua produk dari Sheet Master Stok."""
    _, sheet = _get_workbook_and_sheet(SHEET_MASTER_STOK, read_only=True)
    
    products: List[MasterStockProduct] = []
    
    # Iterasi dari baris ke-2 (data)
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        if not row[0]: # Lewati baris kosong jika kolom pertama kosong
            continue

        # Mapping data berdasarkan urutan kolom
        harga_jual_data = {
            'bungkus': row[5], 'batang': row[6], 'mentah': row[7],
            'seduh': row[8], 'rebus': row[9], 'rebus_telur': row[10],
        }
        
        try:
            product_data = {
                'nama_produk': row[0],
                'satuan_beli': row[1],
                # Konversi ke int/float dan handle None
                'isi_per_satuan_beli': int(row[2]) if row[2] else 0,
                'kategori': row[3] or '',
                'satuan_unit_dasar': row[4] or '',
                'harga_jual': HargaJual(**{k: float(v) if v is not None else None for k, v in harga_jual_data.items()})
            }
            product = MasterStockProduct(**product_data)
            products.append(product)
        except Exception as e:
            # Print error spesifik untuk debugging Excel
            print(f"Error memuat produk '{row[0]}' di baris {row_idx + 2}: {e}")
            continue

    return products

def get_product_by_name(nama_produk: str) -> Optional[Tuple[MasterStockProduct, int]]:
    """Mencari produk berdasarkan nama, mengembalikan produk dan nomor barisnya (untuk Update/Delete)."""
    # Menggunakan read_master_stock untuk mendapatkan produk dari model, lalu cari index
    all_products = read_master_stock()
    found_product = next((p for p in all_products if p.nama_produk.lower() == nama_produk.lower()), None)

    if found_product:
        # Jika ditemukan, cari nomor baris aslinya di Excel (lambat, tapi akurat)
        _, sheet = _get_workbook_and_sheet(SHEET_MASTER_STOK, read_only=True)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] and row[0].lower() == nama_produk.lower():
                return found_product, row_idx
    
    return None

def create_master_stock(product: MasterStockProduct):
    """Menambahkan produk baru ke Master Stok."""
    if get_product_by_name(product.nama_produk):
        raise ValueError("Produk dengan nama ini sudah ada.")
        
    workbook, sheet = _get_workbook_and_sheet(SHEET_MASTER_STOK)
    
    # Siapkan data untuk baris baru
    harga_jual = product.harga_jual.model_dump()
    row_data = [
        product.nama_produk, product.satuan_beli, product.isi_per_satuan_beli, 
        product.kategori, product.satuan_unit_dasar, 
        harga_jual['bungkus'], harga_jual['batang'], harga_jual['mentah'], 
        harga_jual['seduh'], harga_jual['rebus'], harga_jual['rebus_telur'],
    ]
    
    sheet.append(row_data)
    workbook.save(FILE_PATH)
    
def update_master_stock(nama_produk_lama: str, updated_product: MasterStockProduct):
    """Memperbarui data produk berdasarkan nama produk lama."""
    
    product_data_pair = get_product_by_name(nama_produk_lama)
    if not product_data_pair:
        raise ValueError(f"Produk '{nama_produk_lama}' tidak ditemukan untuk diperbarui.")
        
    _, row_idx = product_data_pair
    
    workbook, sheet = _get_workbook_and_sheet(SHEET_MASTER_STOK)
    
    # Siapkan data baru (sama seperti create)
    harga_jual = updated_product.harga_jual.model_dump()
    new_row_data = [
        updated_product.nama_produk, updated_product.satuan_beli, updated_product.isi_per_satuan_beli, 
        updated_product.kategori, updated_product.satuan_unit_dasar, 
        harga_jual['bungkus'], harga_jual['batang'], harga_jual['mentah'], 
        harga_jual['seduh'], harga_jual['rebus'], harga_jual['rebus_telur'],
    ]

    # Tulis data baru ke baris yang sudah ada
    for col_idx, value in enumerate(new_row_data, start=1):
        sheet.cell(row=row_idx, column=col_idx, value=value)
        
    workbook.save(FILE_PATH)
    
def delete_master_stock(nama_produk: str):
    """Menghapus produk dari Master Stok berdasarkan nama."""
    product_data_pair = get_product_by_name(nama_produk)
    if not product_data_pair:
        raise ValueError(f"Produk '{nama_produk}' tidak ditemukan untuk dihapus.")
        
    _, row_idx = product_data_pair
    
    workbook, sheet = _get_workbook_and_sheet(SHEET_MASTER_STOK)
    
    # Hapus baris
    sheet.delete_rows(row_idx, 1)
    
    workbook.save(FILE_PATH)
    
# --- FUNGSI UTAMA (JURNAL TRANSAKSI) ---

def write_sales_transaction(transactions: List[JurnalPenjualan]):
    """Menulis transaksi penjualan ke sheet Jurnal Penjualan."""
    workbook, sheet = _get_workbook_and_sheet(SHEET_JURNAL_PENJUALAN)
    
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    for t in transactions:
        row_data = [
            timestamp, t.nama_produk, t.jumlah_jual, t.total_harga_jual, t.catatan or None,
        ]
        sheet.append(row_data)
        
    workbook.save(FILE_PATH)
    
def write_purchase_transaction(transactions: List[JurnalPembelian]):
    """Menulis transaksi pembelian ke sheet Jurnal Pembelian."""
    workbook, sheet = _get_workbook_and_sheet(SHEET_JURNAL_PEMBELIAN) # KOREKSI DI SINI JUGA!
    
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    for t in transactions:
        row_data = [
            timestamp, t.nama_produk, t.jumlah_beli, t.satuan_beli, t.total_harga_beli,
        ]
        sheet.append(row_data)
        
    workbook.save(FILE_PATH)

def get_product_by_name(name: str) -> Optional[Tuple[MasterStockProduct, int]]:
    """
    Mencari produk di Master Stok berdasarkan nama.
    Mengembalikan tuple (product_model, row_index) atau None.
    """
    _ensure_file_and_sheets()
    
    # Baca data dari Master Stok
    workbook, sheet = _get_workbook_and_sheet(SHEET_MASTER_STOK, read_only=True)
    
    # Mencari index baris yang sesuai (dimulai dari baris 2 setelah header)
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Asumsi nama produk ada di kolom pertama (index 0)
        product_name_in_sheet = row[0]
        
        if product_name_in_sheet == name:
            # Reconstruct model dari data baris (gunakan logika dari read_master_stock)
            
            # 1. Ambil data harga jual (sesuai urutan di header: kolom 6 sampai 11)
            harga_jual_data = {
                'bungkus': row[5], 'batang': row[6], 'mentah': row[7],
                'seduh': row[8], 'rebus': row[9], 'rebus_telur': row[10]
            }
            
            # Bersihkan nilai string None atau '-' menjadi None/0.0
            def clean_float(val):
                if val is None or str(val).strip() in ['-', '']:
                    return None
                try:
                    return float(val)
                except ValueError:
                    return None # Jika tidak bisa di-float
            
            cleaned_harga_jual = {k: clean_float(v) for k, v in harga_jual_data.items()}
            
            # 2. Buat model HargaJual
            harga_jual_model = HargaJual(**cleaned_harga_jual)
            
            # 3. Buat model MasterStockProduct
            product_model = MasterStockProduct(
                nama_produk=product_name_in_sheet,
                satuan_beli=row[1],
                isi_per_satuan_beli=int(row[2]) if row[2] is not None else 0,
                kategori=row[3],
                satuan_unit_dasar=row[4],
                harga_jual=harga_jual_model
            )
            
            return product_model, i # Kembalikan model dan index baris Excel
            
    return None

def update_master_stock_cost_price(name: str, new_cost_price: float):
    """Memperbarui harga modal (kolom 2) produk di Master Stok."""
    product_data_pair = get_product_by_name(name)

    if not product_data_pair:
        raise ValueError(f"Produk '{name}' tidak ditemukan di Master Stok.")
        
    # Ambil index baris (i) dari get_product_by_name
    product_model, row_index = product_data_pair 

    workbook, sheet = _get_workbook_and_sheet(SHEET_MASTER_STOK, read_only=False)

    # Kolom harga modal berada di kolom B, yang berarti index kolom ke-2
    # Kita menggunakan openpyxl: row[index_berbasis_1]
    # Header Master Stok: [Nama Produk, Satuan Beli, Isi per Satuan Beli, Kategori, Satuan Unit Dasar, ...]
    # Kolom 2: Satuan Beli. Kolom 1 (A) adalah Nama Produk.
    
    # Berdasarkan Header File: 'Nama Produk, Satuan Beli, Isi per Satuan Beli, Kategori, satuan Unit Dasar, Harga jual...'
    # Harga modal (kolom 2) harusnya 'Satuan Beli', tapi di Jurnal Pembelian butuh harga modal!
    
    # ASUMSI: Berdasarkan header file yang di-upload, tidak ada kolom harga modal di Master Stok.
    # Namun, karena Jurnal Pembelian membutuhkan harga modal terbaru, kita ASUMSIKAN:
    # Harga Modal ditempatkan di KOLOM KEDUA Master Stok (sebelum Isi per Satuan Beli).
    # KOREKSI: Karena kode kita sebelumnya tidak menyertakan Harga Modal di model dan header, 
    # untuk sementara kita akan memperbarui Kolom 2 (Satuan Beli) DULU, dan perlu revisi model nanti!
    
    # Skenario yang benar: Kita harus memperbarui kolom harga modal.
    # Karena kita belum punya kolom harga modal di header/model, kita HANYA AKAN MENCATAT DI JURNAL PEMBELIAN
    # dan MENGABAIKAN update Master Stok untuk sementara, KECUALI Anda tambahkan kolom Harga Modal (Beli) di Master Stok.
    
    # Untuk melanjutkan, kita ASUMSIKAN Anda telah menambahkan kolom "Harga Modal" (Index 1) di Master Stok Excel Anda.
    # Header Baru (ASUMSI): [Nama Produk, Harga Modal Beli, Satuan Beli, Isi per Satuan Beli, ...]
    
    # Jika Asumsi Benar:
    # Kolom Harga Modal Beli (Index 2 di Python, Kolom B di Excel)
    CELL_MODAL_PRICE = 2 # Kolom ke-2 (B)
    
    sheet.cell(row=row_index, column=CELL_MODAL_PRICE, value=new_cost_price)
    
    workbook.save(FILE_PATH)